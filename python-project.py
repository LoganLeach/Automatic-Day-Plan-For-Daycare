import openpyxl, random, datetime, os, docx

# This function asks the user to type what word doc they would like the day plan to print to, if they get the incorrect name they get a error
def validFileName(pro):
    fileName = input(pro)
    while not os.path.exists(fileName):
        print(f"File '{fileName}' was not found please try typing the file name again (make sure to add the correct file extension name).")
        fileName = input(pro)
    return fileName

try:
    # Get the Word document file name
    docwordName = validFileName("\nEnter the Word document file name you would like the day plan on: ")

    # Get the Excel spreadsheet file name
    excelSheetName = validFileName("\nEnter the Excel spreadsheet file name you have your schedule on: ")
    print(f"\nThe word and excel file was found!\n")
except Exception as error:
    print(f"An error occurred: {error}")
    
doc = docx.Document()

# This function loads a spreadsheet called daycaree prints the information from it onto a word doc (what kids and there arrival time), 
# it also prints a title, todays date
def arriveTime(doc):
    workbook = openpyxl.load_workbook(excelSheetName, read_only=True)
    worksheet = workbook.active

    doc.add_paragraph("Happy Friendly DayCare Day Plan:")
    time = datetime.datetime.today().strftime('%m/%d/%Y')
    doc.add_paragraph(f"Todays Date is: {time}")
    doc.add_paragraph("--------------------------------------------------------------------------------------")
    doc.add_paragraph("Today the following kids are arriving:")

    maxNameLength = 20 #This sets the name length so if you accidently make a really long name and dont realize it will cut it off at 20 so the doc doesnt look messy and you see your naming error
    for row in worksheet.iter_rows(min_row=2, max_col=2):
        time = row[0].value
        name = row[1].value[:maxNameLength]

        doc.add_paragraph(f"{name} will be arriving at {time} today!")

arriveTime(doc)

doc.add_paragraph("--------------------------------------------------------------------------------------")

# In this function it will ask the user if there having breakfast, lunch, or snack today, if yes then it will print that choice 
# if no then it wont print it, if you happen to spell yes/no wrong it will give you a error and make you retry. It also has 
# 3 lists of food options that are later used to print onto the word doc
def mealPlan():
    breakfastChoice = ["Banana with granola bar", "Fruit Loops", "Oatmeal", "Chocolate chip muffins", "Pancakes", "Waffles", "Cherrios", "Special K", "Breakfast sausage", "Boiled eggs", "Scrambled egss", "Poached eggs", "Toast", "English muffins"]
    lunchChoice = ["Hotdogs", "Salad", "Kraft dinner with corn", "Pizza", "Buns with cucumbers", "Spagetti and meatballs", "Ham Sandwhich", "Roast beef sandwhich", "Mac and Cheese", "Pork Ribitte", ]
    snackChoice = ["Bear paws", "Oreos with milk", "Pretzels", "Cereal mix", "Chips", "Nutigrain bars", "Greek yogurt", "PBJ sandwhich", "Peanut butter and crackers", "Smoothies"]

    doc.add_paragraph("Here's today's meal plan:")
    
    ifBreakfast = input("Are you making breakfast today? (yes/no): ").lower()
    while ifBreakfast not in ['yes', 'no']:
        print("Sorry that is the inncorrect input. Please enter 'yes' or 'no'.")
        ifBreakfast = input("Are you making breakfast today? (yes/no): ").lower()
    
    if ifBreakfast == 'yes':
        breakfast = random.choice(breakfastChoice)
        doc.add_paragraph("The breakfast today will be: " + breakfast)
    
    IfLunch = input("Are you making lunch today? (yes/no): ").lower()
    while IfLunch not in ['yes', 'no']:
        print("Sorry that is the inncorrect input. Please enter 'yes' or 'no'.")
        IfLunch = input("Are you making lunch today? (yes/no): ").lower()
    
    if IfLunch == 'yes':
        lunch = random.choice(lunchChoice)
        doc.add_paragraph("The lunch today will be: " + lunch)
    
    IfSnack = input("Are you making snack today? (yes/no): ").lower()
    while IfSnack not in ['yes', 'no']:
        print("Sorry that is the inncorrect input. Please enter 'yes' or 'no'.")
        IfSnack = input("Are you making snack today? (yes/no): ").lower()
    
    if IfSnack == 'yes':
        snack = random.choice(snackChoice)
        doc.add_paragraph("The snack today will be: " + snack)

mealPlan()

# This function asks the user how many activities they would like to today, depending on the answer is how many it will print out. 
# It prints it out because of the activities lists and the for loop that loops over the chosenActivites variable that randomizes 
# whats in the list
def activitiesTodo():
    activities = ["Sports", "Nature walks", "outside time", "board games", "hide and seek", "Park", "Freeze tag", "I spy with my little eye", "swimmming"]

    howManyActivities = int(input("How many activities would you like to do today? "))
    activitieChosen = random.sample(activities, howManyActivities)

    doc.add_paragraph("--------------------------------------------------------------------------------------")

    doc.add_paragraph("Here are today's planned activities:")
    for activity in activitieChosen:
        doc.add_paragraph("- " + activity)

activitiesTodo()

# This function loads a spreadsheet called birthdays then loops over it checking the date for any birthdays, if there is a 
# birthday close (7 days) it will let the user know so they have time to prepare
def birthdays():
    workbook = openpyxl.load_workbook('birthdays.xlsx', read_only=True)
    sheet = workbook.active

    today = datetime.datetime.today()
    todayTime = today.strftime('%m-%d')

    birthdaysupcoming = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[0]
        birthday = row[1]
        bday = birthday.strftime('%m-%d')
        if bday == todayTime:
            doc.add_paragraph("--------------------------------------------------------------------------------------")
            doc.add_paragraph("It is " + name + "'s birthday today, make sure to wish them a happy birthday!")
        elif (birthday - today).days <= 7 and (birthday - today).days > 0: # lets the user know its someones birthday in 7 days
            birthdaysupcoming.append((name, birthday))

    if birthdaysupcoming:
        doc.add_paragraph("--------------------------------------------------------------------------------------")
        doc.add_paragraph("Upcoming Birthdays:")
        for name, birthday in birthdaysupcoming:
            days_until_bday = (birthday - today).days
            doc.add_paragraph(f"{name}'s birthday is coming up in {days_until_bday} days!")
    else:
        doc.add_paragraph("--------------------------------------------------------------------------------------")
        doc.add_paragraph("No birthdays coming up in the next week.")

birthdays()

# This saves the docx file
doc.save('todays_schedule.docx')
#This starts the docx file so the program can locate it in the file system 
os.system("start todays_schedule.docx")